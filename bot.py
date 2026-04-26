import os
import time
import logging
import warnings
import traceback
from datetime import datetime, timedelta

import telebot
import pyodbc
import pandas as pd
from telebot.types import ReplyKeyboardMarkup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from PIL import Image, ImageDraw, ImageFont

warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")

# ================= CONFIG =================
TOKEN = "8573812581:AAEpYmJD8ocfyiOwWuWypN2PM6e91J0mhQk"   # <-- BU YERGA YANGI TOKEN YOZING
SERVER = "."
DATABASE = "NGLOBAL"
ADMIN_ID = 1405847283
BOT_NAME = "firma_bot"
LOGO_PATH = "logo.png"

logging.basicConfig(level=logging.INFO)
bot = telebot.TeleBot(TOKEN)

# ================= STATES =================
user_states = {}
STATE_ADD_USER = "add_user"
STATE_DELETE_USER = "delete_user"

# ================= SQL CONNECTION =================
conn = pyodbc.connect(
    "DRIVER={SQL Server Native Client 11.0};"
    f"SERVER={SERVER};"
    f"DATABASE={DATABASE};"
    "Trusted_Connection=yes;"
)
conn.autocommit = True
cursor = conn.cursor()

print("✅ SQL ulandi")
print("BOT FILE:", os.path.abspath(__file__))


# ================= ACCESS =================
def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


def is_allowed(user_id: int) -> bool:
    if is_admin(user_id):
        return True

    try:
        cursor.execute(
            "SELECT 1 FROM bot_users WHERE telegram_id = ? AND bot_name = ?",
            user_id, BOT_NAME
        )
        return cursor.fetchone() is not None
    except Exception as e:
        print("USER CHECK XATO:", e)
        return False


def user_has_all_producers(user_id: int) -> bool:
    if is_admin(user_id):
        return True

    try:
        cursor.execute("""
            SELECT ISNULL(all_producers, 0)
            FROM bot_users
            WHERE telegram_id = ? AND bot_name = ?
        """, user_id, BOT_NAME)
        row = cursor.fetchone()
        return bool(row[0]) if row else False
    except Exception as e:
        print("GET ALL_PRODUCERS XATO:", e)
        return False


def get_user_producer_ids(user_id: int):
    if is_admin(user_id):
        return []

    try:
        cursor.execute("""
            SELECT producer_id
            FROM bot_user_producers
            WHERE telegram_id = ? AND bot_name = ?
            ORDER BY producer_id
        """, user_id, BOT_NAME)
        rows = cursor.fetchall()
        return [int(r[0]) for r in rows]
    except Exception as e:
        print("GET USER PRODUCERS XATO:", e)
        return []


def get_producer_name_by_id(producer_id: int):
    try:
        cursor.execute("SELECT NAME FROM S_PRODUCER WHERE ID = ?", producer_id)
        row = cursor.fetchone()
        return row[0].strip() if row and row[0] else None
    except Exception as e:
        print("GET PRODUCER NAME XATO:", e)
        return None


def get_user_producer_names(user_id: int):
    ids = get_user_producer_ids(user_id)
    names = []

    for pid in ids:
        pname = get_producer_name_by_id(pid)
        if pname:
            names.append((pid, pname))

    return names


def get_access_denied_text(user_id: int) -> str:
    return (
        "⛔ Sizda botdan foydalanish uchun dostup yo'q.\n"
        "⛔ У вас нет доступа к боту.\n\n"
        f"Sizning ID / Ваш ID: <code>{user_id}</code>\n\n"
        "Dostup olish uchun admin bilan bog'laning va yuqoridagi ID ni yuboring.\n"
        "Для получения доступа свяжитесь с администратором и отправьте указанный ID."
    )


# ================= MENUS =================
def admin_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row("➕ User qo‘shish", "🗑 User o‘chirish")
    kb.row("📋 Userlar")
    return kb


def user_menu(producer_names):
    kb = ReplyKeyboardMarkup(resize_keyboard=True)

    row = []
    for _, pname in producer_names:
        row.append(f"🏭 {pname}")
        if len(row) == 2:
            kb.row(*row)
            row = []

    if row:
        kb.row(*row)

    return kb


# ================= PDF FONT =================
def register_pdf_fonts():
    font_dir = r"C:\Windows\Fonts"

    regular_font = os.path.join(font_dir, "arial.ttf")
    bold_font = os.path.join(font_dir, "arialbd.ttf")

    if not os.path.exists(regular_font) or not os.path.exists(bold_font):
        raise FileNotFoundError(
            "Arial shrift topilmadi. C:\\Windows\\Fonts ichida arial.ttf va arialbd.ttf bo‘lishi kerak."
        )

    try:
        pdfmetrics.getFont("ArialCustom")
    except Exception:
        pdfmetrics.registerFont(TTFont("ArialCustom", regular_font))

    try:
        pdfmetrics.getFont("ArialCustomBold")
    except Exception:
        pdfmetrics.registerFont(TTFont("ArialCustomBold", bold_font))


# ================= HELPERS =================
def fmt_money(x):
    return f"{float(x):,.2f}".replace(",", " ")


def safe_edit_or_send(chat_id, message_id, text, reply_markup=None):
    try:
        bot.edit_message_text(
            text,
            chat_id=chat_id,
            message_id=message_id
        )
    except Exception:
        bot.send_message(
            chat_id,
            text,
            reply_markup=reply_markup
        )


def build_group_df(image_df: pd.DataFrame) -> pd.DataFrame:
    """
    Image uchun tovar bo‘yicha umumiy itogo.
    Filiallar bo‘yicha qoldiq/summa yig‘iladi.
    Sotuv esa bir xil tovar satrlarida takrorlanib qolmasligi uchun MAX olinadi.
    """
    group_df = (
        image_df.groupby("Наименование товара", as_index=False)
        .agg({
            "Остаток": "sum",
            "Продано за 1 ой": "max",
            "Сумма": "sum"
        })
        .sort_values(by="Наименование товара")
        .reset_index(drop=True)
    )

    total_row = pd.DataFrame([{
        "Наименование товара": "ИТОГО:",
        "Остаток": group_df["Остаток"].sum(),
        "Продано за 1 ой": group_df["Продано за 1 ой"].sum(),
        "Сумма": group_df["Сумма"].sum()
    }])

    group_df = pd.concat([total_row, group_df], ignore_index=True)
    return group_df


def create_itogo_image(group_df: pd.DataFrame, producer_name: str):
    data = group_df.copy()

    data["Наименование товара"] = data["Наименование товара"].astype(str)
    data["Остаток"] = data["Остаток"].astype(float)
    data["Продано за 1 ой"] = data["Продано за 1 ой"].astype(float)
    data["Сумма"] = data["Сумма"].astype(float)

    arial = r"C:\Windows\Fonts\arial.ttf"
    arialbd = r"C:\Windows\Fonts\arialbd.ttf"

    width = 1320
    title_h = 150
    header_h = 48
    row_h = 50
    footer_h = 20

    font_title = ImageFont.truetype(arialbd, 30)
    font_head = ImageFont.truetype(arialbd, 22)
    font_text = ImageFont.truetype(arial, 21)
    font_bold = ImageFont.truetype(arialbd, 21)

    x_name = 20
    x_ost = 760
    x_sale = 920
    x_sum = 1100

    rows_per_image = 25
    file_names = []

    total_rows = len(data)
    chunks = [data.iloc[i:i + rows_per_image].copy() for i in range(0, total_rows, rows_per_image)]

    for part_idx, chunk in enumerate(chunks, start=1):
        height = title_h + header_h + (len(chunk) * row_h) + footer_h
        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        draw.rectangle([0, 0, width, title_h], fill="#E8D7C9")

        if os.path.exists(LOGO_PATH):
            try:
                logo = Image.open(LOGO_PATH).convert("RGBA")
                logo.thumbnail((110, 110))
                logo_x = (width - logo.width) // 2
                logo_y = 8
                img.paste(logo, (logo_x, logo_y), logo)
            except Exception:
                pass

        title_text = (
            f"{producer_name} — ИТОГО ПО ТОВАРАМ ({part_idx}/{len(chunks)})"
            if len(chunks) > 1 else
            f"{producer_name} — ИТОГО ПО ТОВАРАМ"
        )
        title_bbox = draw.textbbox((0, 0), title_text, font=font_title)
        title_w = title_bbox[2] - title_bbox[0]
        draw.text(((width - title_w) // 2, 115), title_text, fill="black", font=font_title)

        y = title_h
        draw.rectangle([0, y, width, y + header_h], fill="#EFEFEF", outline="#999999")
        draw.text((x_name, y + 11), "Наименование товара", fill="black", font=font_head)
        draw.text((x_ost, y + 11), "Остаток", fill="black", font=font_head)
        draw.text((x_sale, y + 11), "1 oyda sotilishi", fill="black", font=font_head)
        draw.text((x_sum, y + 11), "Сумма", fill="black", font=font_head)

        y += header_h

        for i, (_, row) in enumerate(chunk.iterrows()):
            is_total = str(row["Наименование товара"]).strip().upper() in ["ИТОГО:", "ИТОГО"]

            if is_total:
                bg = "#E8D7C9"
                current_font = font_bold
            else:
                bg = "#FAFAFA" if i % 2 == 0 else "white"
                current_font = font_text

            draw.rectangle([0, y, width, y + row_h], fill=bg, outline="#C0C0C0")

            name = str(row["Наименование товара"])
            if len(name) > 42 and not is_total:
                name = name[:42] + "..."

            ost = f'{float(row["Остаток"]):.2f}'
            sale = f'{float(row["Продано за 1 ой"]):.2f}'
            summ = fmt_money(row["Сумма"])

            draw.text((x_name, y + 12), name, fill="black", font=current_font)
            draw.text((x_ost, y + 12), ost, fill="black", font=current_font)
            draw.text((x_sale, y + 12), sale, fill="black", font=current_font)
            draw.text((x_sum, y + 12), summ, fill="black", font=current_font)

            y += row_h

        file_name = f"{BOT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_itogo_{part_idx}.png"
        img.save(file_name)
        file_names.append(file_name)

    return file_names


# ================= EXCEL =================
def style_excel_sheet(ws, producer_name: str):
    top_fill = PatternFill(fill_type="solid", fgColor="E8D7C9")
    header_fill = PatternFill(fill_type="solid", fgColor="E8D7C9")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    big_font = Font(name="Calibri", size=18, bold=True)
    header_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    total_font = Font(name="Calibri", size=11, bold=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F2")
    ws["A1"] = f"Производитель:    {producer_name.upper()}"
    ws["A1"].font = big_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = top_fill

    for r in range(1, 3):
        for c in range(1, 7):
            cell = ws.cell(r, c)
            cell.fill = top_fill
            cell.border = border

    for c in range(1, 7):
        cell = ws.cell(3, c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border

    for r in range(4, ws.max_row + 1):
        for c in range(1, 7):
            cell = ws.cell(r, c)
            cell.border = border
            cell.fill = white_fill
            cell.font = total_font if r == ws.max_row else normal_font

        for c in range(1, 3):
            ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="center")

        for c in range(3, 7):
            ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(r, c).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A4"


def make_excel(export_df: pd.DataFrame, producer_name: str) -> str:
    filename = f"{BOT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Itogoda sotuv dubl bo‘lmasligi uchun tovar bo‘yicha MAX
    total_sale = (
        export_df.groupby("Наименование товара", as_index=False)["Продано за 1 ой"]
        .max()["Продано за 1 ой"]
        .sum()
    )

    total_row = pd.DataFrame([{
        "Наименование товара": "ИТОГО:",
        "Отдел": "",
        "Остаток": export_df["Остаток"].sum(),
        "Цена приход": "",
        "Продано за 1 ой": total_sale,
        "Сумма": export_df["Сумма"].sum()
    }])

    export_ready = pd.concat([export_df, total_row], ignore_index=True)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        export_ready.to_excel(writer, index=False, startrow=2, sheet_name="Детально")

    wb = load_workbook(filename)
    style_excel_sheet(wb["Детально"], producer_name)
    wb.save(filename)

    return filename


# ================= PDF =================
def make_pdf(export_df: pd.DataFrame, producer_name: str) -> str:
    register_pdf_fonts()

    filename = f"{BOT_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    total_ost = float(export_df["Остаток"].sum())
    total_sale = float(
        export_df.groupby("Наименование товара", as_index=False)["Продано за 1 ой"]
        .max()["Продано за 1 ой"]
        .sum()
    )
    total_sum = float(export_df["Сумма"].sum())

    pdf_df = export_df.copy().fillna("")
    pdf_df["Наименование товара"] = pdf_df["Наименование товара"].apply(
        lambda x: (str(x)[:28] + "...") if len(str(x)) > 28 else str(x)
    )
    pdf_df["Отдел"] = pdf_df["Отдел"].apply(
        lambda x: (str(x)[:16] + "...") if len(str(x)) > 16 else str(x)
    )

    for col in ["Остаток", "Цена приход", "Продано за 1 ой", "Сумма"]:
        pdf_df[col] = pdf_df[col].apply(
            lambda x: fmt_money(x) if str(x).strip() not in ["", "nan"] else ""
        )

    total_row = pd.DataFrame([{
        "Наименование товара": "ИТОГО:",
        "Отдел": "",
        "Остаток": fmt_money(total_ost),
        "Цена приход": "",
        "Продано за 1 ой": fmt_money(total_sale),
        "Сумма": fmt_money(total_sum)
    }])

    pdf_df = pd.concat([pdf_df, total_row], ignore_index=True)

    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=4,
        leftMargin=4,
        topMargin=4,
        bottomMargin=4
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="TitleCenter",
        parent=styles["Normal"],
        fontName="ArialCustomBold",
        fontSize=13,
        leading=15,
        alignment=1,
        spaceAfter=1
    )

    subtitle_style = ParagraphStyle(
        name="SubTitleCenter",
        parent=styles["Normal"],
        fontName="ArialCustom",
        fontSize=6,
        leading=7,
        alignment=1,
        textColor=colors.HexColor("#666666")
    )

    info_style = ParagraphStyle(
        name="InfoStyle",
        parent=styles["Normal"],
        fontName="ArialCustomBold",
        fontSize=7,
        leading=8,
        alignment=1
    )

    footer_style = ParagraphStyle(
        name="FooterStyle",
        parent=styles["Normal"],
        fontName="ArialCustom",
        fontSize=6,
        leading=7,
        alignment=1,
        textColor=colors.HexColor("#666666")
    )

    elements = []

    if os.path.exists(LOGO_PATH):
        logo = RLImage(LOGO_PATH, width=10 * mm, height=10 * mm)
        logo.hAlign = "CENTER"
        elements.append(logo)
        elements.append(Spacer(1, 1))

    elements.append(Paragraph(f"ПРОИЗВОДИТЕЛЬ: {producer_name.upper()}", title_style))
    elements.append(Paragraph("Остаток товаров по отделам и продажи за 1 oy", subtitle_style))
    elements.append(Spacer(1, 2))

    summary_data = [[
        Paragraph(f"ПОЗИЦИЙ:<br/>{len(export_df)}", info_style),
        Paragraph(f"ОСТАТОК ИТОГО:<br/>{fmt_money(total_ost)}", info_style),
        Paragraph(f"ПРОДАНО ЗА 1 ОЙ:<br/>{fmt_money(total_sale)}", info_style),
        Paragraph(f"СУММА ИТОГО:<br/>{fmt_money(total_sum)}", info_style),
    ]]

    summary_table = Table(summary_data, colWidths=[47 * mm, 47 * mm, 47 * mm, 47 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4E7DA")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#B8A28D")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B8A28D")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 4))

    detail_box = Table([["ДЕТАЛЬНЫЙ ОТЧЕТ"]], colWidths=[190 * mm])
    detail_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E8D7C9")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("FONTNAME", (0, 0), (-1, -1), "ArialCustomBold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#B8A28D")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(detail_box)
    elements.append(Spacer(1, 3))

    detail_data = [list(pdf_df.columns)] + pdf_df.values.tolist()
    detail_table = Table(
        detail_data,
        repeatRows=1,
        colWidths=[44 * mm, 28 * mm, 20 * mm, 24 * mm, 24 * mm, 30 * mm]
    )

    detail_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFEFEF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "ArialCustomBold"),
        ("FONTSIZE", (0, 0), (-1, 0), 6.0),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "ArialCustom"),
        ("FONTSIZE", (0, 1), (-1, -1), 5.2),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]

    for row_num in range(1, len(detail_data)):
        row_name = str(detail_data[row_num][0]).strip().upper()
        if row_name in ["ИТОГО:", "ИТОГО"]:
            detail_style.append(("BACKGROUND", (0, row_num), (-1, row_num), colors.HexColor("#E8D7C9")))
            detail_style.append(("FONTNAME", (0, row_num), (-1, row_num), "ArialCustomBold"))
        else:
            detail_style.append((
                "BACKGROUND",
                (0, row_num),
                (-1, row_num),
                colors.HexColor("#FAFAFA") if row_num % 2 == 0 else colors.white
            ))

    detail_table.setStyle(TableStyle(detail_style))
    elements.append(detail_table)
    elements.append(Spacer(1, 2))

    elements.append(Paragraph(
        f"Hisobot sanasi: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        footer_style
    ))

    doc.build(elements)
    return filename


# ================= DATA HELPERS =================
def get_producer_id_by_name_or_id(text: str):
    try:
        return int(text)
    except Exception:
        pass

    cursor.execute("""
        SELECT TOP 1 ID, NAME
        FROM S_PRODUCER
        WHERE LTRIM(RTRIM(NAME)) COLLATE Cyrillic_General_CI_AI
              LIKE LTRIM(RTRIM(?)) + '%'
        ORDER BY NAME
    """, text)
    row = cursor.fetchone()
    return int(row[0]) if row else None


def get_sales_30_from_history_q(producer_id: int, conn) -> pd.DataFrame:
    print("DEBUG: HISTORY_Q VERSION IS WORKING")

    cur = conn.cursor()

    today = datetime.now().date()
    date_begin = datetime.combine(today - timedelta(days=30), datetime.min.time())
    date_end = datetime.combine(today, datetime.max.time())

    print("DATEBEGIN =", date_begin)
    print("DATEEND   =", date_end)

    cur.execute("""
        SELECT ID, NAME
        FROM GOOD
        WHERE PRODUCER = ?
        ORDER BY NAME
    """, producer_id)

    goods = cur.fetchall()
    rows = []

    for g in goods:
        good_id = int(g[0])
        good_name = str(g[1]).strip() if g[1] else ""

        try:
            df_hist = pd.read_sql("""
                EXEC dbo.HISTORY_Q
                    @INCOMELN = 0,
                    @GOOD = ?,
                    @FUN = 5,
                    @OTDEL = 0,
                    @DATEBEGIN = ?,
                    @DATEEND = ?
            """, conn, params=[good_id, date_begin, date_end])

            total = float(df_hist["KOL"].fillna(0).sum()) if not df_hist.empty else 0.0

            rows.append({
                "GOOD_ID": good_id,
                "Наименование товара": good_name,
                "Продано за 1 ой": total
            })

        except Exception as e:
            print(f"HISTORY_Q XATO good_id={good_id}: {e}")
            rows.append({
                "GOOD_ID": good_id,
                "Наименование товара": good_name,
                "Продано за 1 ой": 0.0
            })

    if not rows:
        return pd.DataFrame(columns=["GOOD_ID", "Наименование товара", "Продано за 1 ой"])

    sales_df = pd.DataFrame(rows)

    sales_df = sales_df.groupby(
        ["GOOD_ID", "Наименование товара"], as_index=False
    )["Продано за 1 ой"].sum()

    print("SALES_DF HEAD:")
    try:
        print(sales_df.head(20).to_string(index=False))
    except Exception:
        print(sales_df.head(20))

    return sales_df


# ================= USER SAVE HELPERS =================
def save_user_access(user_id: int, producer_raw: str):
    cursor.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM bot_users
            WHERE telegram_id = ? AND bot_name = ?
        )
        INSERT INTO bot_users (telegram_id, bot_name, all_producers)
        VALUES (?, ?, 0)
    """, user_id, BOT_NAME, user_id, BOT_NAME)

    cursor.execute("""
        DELETE FROM bot_user_producers
        WHERE telegram_id = ? AND bot_name = ?
    """, user_id, BOT_NAME)

    if producer_raw.strip().lower() == "all":
        cursor.execute("""
            UPDATE bot_users
            SET all_producers = 1
            WHERE telegram_id = ? AND bot_name = ?
        """, user_id, BOT_NAME)
        return None, "all"

    producer_ids = [int(x.strip()) for x in producer_raw.split(",") if x.strip()]

    cursor.execute("""
        UPDATE bot_users
        SET all_producers = 0
        WHERE telegram_id = ? AND bot_name = ?
    """, user_id, BOT_NAME)

    for pid in producer_ids:
        cursor.execute("""
            IF NOT EXISTS (
                SELECT 1 FROM bot_user_producers
                WHERE telegram_id = ? AND bot_name = ? AND producer_id = ?
            )
            INSERT INTO bot_user_producers (telegram_id, bot_name, producer_id)
            VALUES (?, ?, ?)
        """, user_id, BOT_NAME, pid, user_id, BOT_NAME, pid)

    return producer_ids, "list"


# ================= BUTTONS =================
@bot.message_handler(func=lambda m: m.text == "➕ User qo‘shish")
def add_user_button(message):
    if not is_admin(message.from_user.id):
        return
    user_states[message.from_user.id] = STATE_ADD_USER
    bot.send_message(
        message.chat.id,
        "Format yuboring:\n123456789 315,316,400\nyoki\n123456789 all",
        reply_markup=admin_menu()
    )


@bot.message_handler(func=lambda m: m.text == "🗑 User o‘chirish")
def del_user_button(message):
    if not is_admin(message.from_user.id):
        return
    user_states[message.from_user.id] = STATE_DELETE_USER
    bot.send_message(
        message.chat.id,
        "O‘chiriladigan user ID ni yuboring:",
        reply_markup=admin_menu()
    )


@bot.message_handler(func=lambda m: m.text == "📋 Userlar")
def users_button(message):
    if not is_admin(message.from_user.id):
        return

    try:
        cursor.execute("""
            SELECT telegram_id, ISNULL(all_producers, 0)
            FROM bot_users
            WHERE bot_name = ?
            ORDER BY telegram_id
        """, BOT_NAME)
        rows = cursor.fetchall()

        if not rows:
            bot.send_message(message.chat.id, f"{BOT_NAME} uchun userlar yo‘q", reply_markup=admin_menu())
            return

        msg = f"{BOT_NAME} userlari:\n\n"

        for r in rows:
            telegram_id = r[0]
            all_producers = bool(r[1])

            if all_producers:
                msg += f"• {telegram_id} — 🔓 BARCHASI\n"
            else:
                ids = get_user_producer_ids(telegram_id)
                if not ids:
                    msg += f"• {telegram_id} — producer biriktirilmagan\n"
                else:
                    names = []
                    for pid in ids:
                        pname = get_producer_name_by_id(pid)
                        names.append(f"{pid}:{pname if pname else 'topilmadi'}")
                    msg += f"• {telegram_id} — {', '.join(names)}\n"

        bot.send_message(message.chat.id, msg, reply_markup=admin_menu())

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Xatolik: {e}", reply_markup=admin_menu())


# ================= COMMANDS =================
@bot.message_handler(commands=["adduser"])
def add_user_command(message):
    if not is_admin(message.from_user.id):
        return

    try:
        parts = message.text.split(maxsplit=2)

        if len(parts) != 3:
            bot.reply_to(message, "❌ Format:\n/adduser 123456789 315,316,400\nYoki: /adduser 123456789 all")
            return

        user_id = int(parts[1])
        producer_raw = parts[2].strip()

        result_ids, mode = save_user_access(user_id, producer_raw)

        if mode == "all":
            bot.reply_to(message, f"✅ User qo‘shildi: {user_id}\n🔓 BARCHA PRODUCERLAR")
            return

        lines = []
        for pid in result_ids:
            pname = get_producer_name_by_id(pid)
            lines.append(f"• {pid} — {pname if pname else 'topilmadi'}")

        bot.reply_to(
            message,
            f"✅ User qo‘shildi: {user_id}\n🏭 Producerlar:\n" + "\n".join(lines)
        )

    except Exception as e:
        bot.reply_to(message, f"❌ Xatolik: {e}")


@bot.message_handler(commands=["deluser"])
def del_user_command(message):
    if not is_admin(message.from_user.id):
        return

    try:
        parts = message.text.split()
        if len(parts) != 2:
            bot.reply_to(message, "❌ Format: /deluser 123456789")
            return

        user_id = int(parts[1])

        if user_id == ADMIN_ID:
            bot.reply_to(message, "❌ Adminni o‘chirib bo‘lmaydi")
            return

        cursor.execute(
            "DELETE FROM bot_user_producers WHERE telegram_id = ? AND bot_name = ?",
            user_id, BOT_NAME
        )
        cursor.execute(
            "DELETE FROM bot_users WHERE telegram_id = ? AND bot_name = ?",
            user_id, BOT_NAME
        )

        bot.reply_to(message, f"🗑 User o‘chirildi: {user_id}")

    except Exception as e:
        bot.reply_to(message, f"❌ Xatolik: {e}")


@bot.message_handler(commands=["users"])
def users_list_command(message):
    if not is_admin(message.from_user.id):
        return

    try:
        cursor.execute("""
            SELECT telegram_id, ISNULL(all_producers, 0)
            FROM bot_users
            WHERE bot_name = ?
            ORDER BY telegram_id
        """, BOT_NAME)
        rows = cursor.fetchall()

        if not rows:
            bot.reply_to(message, f"{BOT_NAME} uchun userlar yo‘q")
            return

        msg = f"{BOT_NAME} userlari:\n\n"

        for r in rows:
            telegram_id = r[0]
            all_producers = bool(r[1])

            if all_producers:
                msg += f"• {telegram_id} — 🔓 BARCHASI\n"
            else:
                ids = get_user_producer_ids(telegram_id)
                if not ids:
                    msg += f"• {telegram_id} — producer biriktirilmagan\n"
                else:
                    names = []
                    for pid in ids:
                        pname = get_producer_name_by_id(pid)
                        names.append(f"{pid}:{pname if pname else 'topilmadi'}")
                    msg += f"• {telegram_id} — {', '.join(names)}\n"

        bot.reply_to(message, msg)

    except Exception as e:
        bot.reply_to(message, f"❌ Xatolik: {e}")


# ================= START =================
@bot.message_handler(commands=["start"])
def start_handler(message):
    user_id = message.from_user.id

    if not is_allowed(user_id):
        bot.send_message(
            message.chat.id,
            get_access_denied_text(user_id),
            parse_mode="HTML"
        )
        return

    if is_admin(user_id):
        bot.send_message(
            message.chat.id,
            f"👑 Admin panel\n🤖 Bot: {BOT_NAME}",
            reply_markup=admin_menu()
        )
        return

    if user_has_all_producers(user_id):
        bot.send_message(
            message.chat.id,
            "✅ Ruxsat bor\n🔓 Siz barcha producerlarni ko‘ra olasiz.\nProducer ID yoki nomini yuboring."
        )
        return

    producer_list = get_user_producer_names(user_id)

    if not producer_list:
        bot.send_message(message.chat.id, "⛔ Sizga producer biriktirilmagan")
        return

    producer_text = "\n".join([f"• {name} (ID: {pid})" for pid, name in producer_list])

    bot.send_message(
        message.chat.id,
        f"✅ Ruxsat bor\n🏭 Sizga biriktirilgan producerlar:\n{producer_text}",
        reply_markup=user_menu(producer_list)
    )


@bot.message_handler(commands=["help"])
def help_handler(message):
    user_id = message.from_user.id

    if not is_allowed(user_id):
        bot.send_message(
            message.chat.id,
            get_access_denied_text(user_id),
            parse_mode="HTML"
        )
        return

    if is_admin(user_id):
        bot.send_message(
            message.chat.id,
            "Admin uchun:\n/adduser 123456789 315,316,400\n/adduser 123456789 all\n/deluser 123456789\n/users",
            reply_markup=admin_menu()
        )
        return

    if user_has_all_producers(user_id):
        bot.send_message(message.chat.id, "🔓 Siz barcha producerlarni ko‘ra olasiz.")
        return

    producer_list = get_user_producer_names(user_id)
    if not producer_list:
        bot.send_message(message.chat.id, "⛔ Sizga producer biriktirilmagan")
        return

    txt = "\n".join([f"• {name} (ID: {pid})" for pid, name in producer_list])
    bot.send_message(
        message.chat.id,
        f"🏭 Sizga biriktirilgan producerlar:\n{txt}",
        reply_markup=user_menu(producer_list)
    )


# ================= MAIN =================
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_message(message):
    user_id = message.from_user.id
    text = message.text.strip()

    if is_admin(user_id) and user_states.get(user_id) == STATE_ADD_USER:
        try:
            parts = text.split(maxsplit=1)

            if len(parts) != 2:
                bot.send_message(
                    message.chat.id,
                    "❌ Format:\n123456789 315,316,400\nyoki\n123456789 all",
                    reply_markup=admin_menu()
                )
                return

            new_user_id = int(parts[0])
            producer_raw = parts[1].strip()

            result_ids, mode = save_user_access(new_user_id, producer_raw)
            user_states.pop(user_id, None)

            if mode == "all":
                bot.send_message(
                    message.chat.id,
                    f"✅ User qo‘shildi: {new_user_id}\n🔓 BARCHA PRODUCERLAR",
                    reply_markup=admin_menu()
                )
            else:
                lines = []
                for pid in result_ids:
                    pname = get_producer_name_by_id(pid)
                    lines.append(f"• {pid} — {pname if pname else 'topilmadi'}")

                bot.send_message(
                    message.chat.id,
                    f"✅ User qo‘shildi: {new_user_id}\n🏭 Producerlar:\n" + "\n".join(lines),
                    reply_markup=admin_menu()
                )

        except Exception as e:
            bot.send_message(
                message.chat.id,
                f"❌ Xatolik: {e}",
                reply_markup=admin_menu()
            )
        return

    if is_admin(user_id) and user_states.get(user_id) == STATE_DELETE_USER:
        try:
            delete_user_id = int(text)

            if delete_user_id == ADMIN_ID:
                bot.send_message(
                    message.chat.id,
                    "❌ Adminni o‘chirib bo‘lmaydi",
                    reply_markup=admin_menu()
                )
                user_states.pop(user_id, None)
                return

            cursor.execute(
                "DELETE FROM bot_user_producers WHERE telegram_id = ? AND bot_name = ?",
                delete_user_id, BOT_NAME
            )
            cursor.execute(
                "DELETE FROM bot_users WHERE telegram_id = ? AND bot_name = ?",
                delete_user_id, BOT_NAME
            )

            user_states.pop(user_id, None)
            bot.send_message(
                message.chat.id,
                f"🗑 User o‘chirildi: {delete_user_id}",
                reply_markup=admin_menu()
            )
        except Exception as e:
            bot.send_message(
                message.chat.id,
                f"❌ Xatolik: {e}",
                reply_markup=admin_menu()
            )
        return

    if not is_allowed(user_id):
        bot.send_message(
            message.chat.id,
            get_access_denied_text(user_id),
            parse_mode="HTML"
        )
        return

    if text in ["➕ User qo‘shish", "🗑 User o‘chirish", "📋 Userlar"]:
        return

    if not is_admin(user_id):
        if user_has_all_producers(user_id):
            search_producer_id = get_producer_id_by_name_or_id(text)
            if not search_producer_id:
                bot.send_message(
                    message.chat.id,
                    "❌ Producer topilmadi.\nID yoki nomini yuboring."
                )
                return
        else:
            producer_list = get_user_producer_names(user_id)
            allowed_map = {f"🏭 {name}": pid for pid, name in producer_list}

            if text not in allowed_map:
                bot.send_message(
                    message.chat.id,
                    "⛔ Siz faqat o‘zingizga biriktirilgan producer tugmalarini bosishingiz mumkin.",
                    reply_markup=user_menu(producer_list)
                )
                return

            search_producer_id = allowed_map[text]

    else:
        search_producer_id = get_producer_id_by_name_or_id(text)
        if not search_producer_id:
            bot.send_message(
                message.chat.id,
                "❌ Producer topilmadi.\nTo‘liq nomini yoki ID ni yuboring.",
                reply_markup=admin_menu()
            )
            return

    wait_msg = bot.send_message(
        message.chat.id,
        "🔎 Izlep atirman...",
        reply_markup=admin_menu() if is_admin(user_id) else None
    )

    try:
        print(f"DEBUG PRODUCER ID = {search_producer_id}")

        # FILIAL KESIMIDAGI QOLDIQ
        query = """
        SELECT
            G.ID AS GOOD_ID,
            G.NAME AS [Наименование товара],
            O.NAME AS [Отдел],
            CAST(SUM(R.OST) AS DECIMAL(18,2)) AS [Остаток],
            CAST(MAX(R.PRICEPRIH) AS DECIMAL(18,2)) AS [Цена приход],
            P.NAME AS [Производитель]
        FROM RESIDUE R
        JOIN GOOD G ON R.GOOD = G.ID
        JOIN S_PRODUCER P ON G.PRODUCER = P.ID
        JOIN ORG O ON R.OTDEL = O.ID
        WHERE P.ID = ?
        GROUP BY G.ID, G.NAME, O.NAME, P.NAME
        HAVING SUM(R.OST) > 0
        ORDER BY G.NAME, O.NAME
        """

        residue_df = pd.read_sql(query, conn, params=[search_producer_id])

        if residue_df.empty:
            safe_edit_or_send(
                wait_msg.chat.id,
                wait_msg.message_id,
                "❌ Hech narsa topilmadi",
                reply_markup=admin_menu() if is_admin(user_id) else None
            )
            return

        print("RESIDUE_DF HEAD:")
        try:
            print(residue_df.head(20).to_string(index=False))
        except Exception:
            print(residue_df.head(20))

        sales_df = get_sales_30_from_history_q(search_producer_id, conn)

        df = residue_df.merge(
            sales_df[["GOOD_ID", "Продано за 1 ой"]],
            on="GOOD_ID",
            how="left"
        )
        df["Продано за 1 ой"] = df["Продано за 1 ой"].fillna(0.0)

        print("MERGED_DF HEAD:")
        try:
            print(df.head(20).to_string(index=False))
        except Exception:
            print(df.head(20))

        producer_name = str(df.iloc[0]["Производитель"]).strip()

        image_df = df[[
            "Наименование товара",
            "Отдел",
            "Остаток",
            "Цена приход",
            "Продано за 1 ой"
        ]].copy()

        image_df["Остаток"] = image_df["Остаток"].astype(float)
        image_df["Цена приход"] = image_df["Цена приход"].astype(float)
        image_df["Продано за 1 ой"] = image_df["Продано за 1 ой"].astype(float)
        image_df["Сумма"] = (image_df["Остаток"] * image_df["Цена приход"]).round(2)

        # IMAGE uchun tovar bo‘yicha itogo
        group_df = build_group_df(image_df)

        # Excel / PDF uchun filial bilan detal hisobot
        export_df = image_df[[
            "Наименование товара",
            "Отдел",
            "Остаток",
            "Цена приход",
            "Продано за 1 ой",
            "Сумма"
        ]].copy()

        excel_file = make_excel(export_df, producer_name)
        pdf_file = make_pdf(export_df, producer_name)
        image_files = create_itogo_image(group_df, producer_name)

        total_sale = float(
            group_df[group_df["Наименование товара"] != "ИТОГО:"]["Продано за 1 ой"].sum()
        )

        safe_edit_or_send(
            wait_msg.chat.id,
            wait_msg.message_id,
            f"✅ Tawildi\n"
            f"🏭 Производитель: {producer_name}\n"
            f"📦 Qatarlar sani: {len(export_df)}\n"
            f"📊 Uliwma qaldiq: {export_df['Остаток'].sum():,.2f}\n"
            f"📉 1 oyda satildi: {total_sale:,.2f}\n"
            f"💰 Сумма итого: {export_df['Сумма'].sum():,.2f}",
            reply_markup=admin_menu() if is_admin(user_id) else None
        )

        for idx, image_file in enumerate(image_files, start=1):
            with open(image_file, "rb") as f:
                bot.send_photo(
                    message.chat.id,
                    f,
                    caption=f"🖼 {producer_name} boyinsha ИТОГО ПО ТОВАРАМ ({idx}/{len(image_files)})"
                )

        with open(excel_file, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption=f"📊 {producer_name} boyinsha Excel tayyar"
            )

        with open(pdf_file, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption=f"📄 {producer_name} boyinsha PDF tayyar"
            )

        for file_path in image_files + [excel_file, pdf_file]:
            try:
                os.remove(file_path)
            except Exception:
                pass

    except Exception as e:
        logging.exception("Xatolik")
        safe_edit_or_send(
            wait_msg.chat.id,
            wait_msg.message_id,
            f"❌ Xatolik: {e}",
            reply_markup=admin_menu() if is_admin(user_id) else None
        )


if __name__ == "__main__":
    print(f"🚀 Bot ishga tushdi: {BOT_NAME}")
    bot.delete_webhook()

    while True:
        try:
            bot.infinity_polling(skip_pending=True, timeout=60, long_polling_timeout=60)
        except Exception as e:
            print(f"❌ Xatolik: {e}")
            traceback.print_exc()
            print("🔄 Qayta ishga tushmoqda...")
            time.sleep(5)
